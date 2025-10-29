#!/usr/bin/env python3
"""
Export dataset statistics to Excel format for visualization and analysis.
Creates formatted Excel files with charts and summary statistics.
"""

import csv
import os
from pathlib import Path
import sys

def export_to_excel(csv_path: Path, excel_path: Path):
    """Export CSV to Excel with formatting and charts."""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Error: Required packages not installed")
        print("Please install: pip install pandas openpyxl")
        return False
    
    if not csv_path.exists():
        print(f"Error: CSV file not found: {csv_path}")
        return False
    
    # Read CSV
    print(f"Reading: {csv_path}")
    df = pd.read_csv(csv_path)
    
    # Create Excel writer
    print(f"Creating Excel file: {excel_path}")
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Data', index=False)
        
        # Create summary statistics
        summary_data = {
            'Metric': [],
            'Value': []
        }
        
        # Add summary stats
        summary_data['Metric'].extend([
            'Total Models',
            'Average Total Tokens',
            'Max Total Tokens',
            'Min Total Tokens',
            'Average Total Size (KB)',
            'Max Total Size (KB)',
            'Min Total Size (KB)',
        ])
        
        summary_data['Value'].extend([
            len(df),
            df['total_tokens'].mean() if 'total_tokens' in df.columns else 'N/A',
            df['total_tokens'].max() if 'total_tokens' in df.columns else 'N/A',
            df['total_tokens'].min() if 'total_tokens' in df.columns else 'N/A',
            df['total_size_kb'].mean() if 'total_size_kb' in df.columns else 'N/A',
            df['total_size_kb'].max() if 'total_size_kb' in df.columns else 'N/A',
            df['total_size_kb'].min() if 'total_size_kb' in df.columns else 'N/A',
        ])
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Create top models sheet
        if 'total_tokens' in df.columns:
            top_models = df.nlargest(20, 'total_tokens')[['model_name', 'total_tokens', 'total_size_kb']]
            top_models.to_excel(writer, sheet_name='Top 20 by Tokens', index=False)
    
    # Format the workbook
    wb = load_workbook(excel_path)
    
    # Format Data sheet
    ws_data = wb['Data']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Format Summary sheet
    ws_summary = wb['Summary']
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # Add chart if top models sheet exists
    if 'Top 20 by Tokens' in wb.sheetnames:
        ws_top = wb['Top 20 by Tokens']
        
        # Format headers
        for cell in ws_top[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        ws_top.column_dimensions['A'].width = 30
        ws_top.column_dimensions['B'].width = 15
        ws_top.column_dimensions['C'].width = 15
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Top 20 Models by Token Count"
        chart.x_axis.title = "Model"
        chart.y_axis.title = "Tokens"
        chart.height = 15
        chart.width = 25
        
        data = Reference(ws_top, min_col=2, min_row=1, max_row=min(21, ws_top.max_row))
        cats = Reference(ws_top, min_col=1, min_row=2, max_row=min(21, ws_top.max_row))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws_top.add_chart(chart, "E2")
    
    wb.save(excel_path)
    print(f"✅ Excel file created successfully!")
    return True


def main():
    """Main function to export statistics to Excel."""
    print("\n" + "="*70)
    print("DATASET STATISTICS - EXCEL EXPORTER")
    print("="*70)
    
    # Check for required packages
    try:
        import pandas
        import openpyxl
    except ImportError:
        print("\n⚠️  Required packages not installed!")
        print("\nPlease install them:")
        print("  pip install pandas openpyxl")
        print("\nOr with the venv:")
        print("  source venv/bin/activate && pip install pandas openpyxl")
        return 1
    
    project_root = Path(__file__).parent
    
    # Get model info from environment
    backend = os.getenv("LLM_BACKEND", "ollama")
    model = os.getenv("LLM_MODEL", "llama3.1:latest")
    
    files_to_export = [
        {
            'name': 'Original Dataset',
            'csv': project_root / 'data' / 'dataset_stats.csv',
            'excel': project_root / 'data' / 'dataset_stats.xlsx'
        },
        {
            'name': 'Generated Dataset',
            'csv': project_root / 'outputs' / f'{backend}_{model}' / 'generated_stats.csv',
            'excel': project_root / 'outputs' / f'{backend}_{model}' / 'generated_stats.xlsx'
        }
    ]
    
    success_count = 0
    for file_info in files_to_export:
        print(f"\n{'='*70}")
        print(f"Exporting: {file_info['name']}")
        print(f"{'='*70}")
        
        if export_to_excel(file_info['csv'], file_info['excel']):
            success_count += 1
            print(f"   Location: {file_info['excel']}")
    
    print(f"\n{'='*70}")
    print(f"Export Complete: {success_count}/{len(files_to_export)} files")
    print(f"{'='*70}")
    
    if success_count > 0:
        print("\n📊 Excel files ready for analysis!")
        print("\nYou can now:")
        print("  1. Open in Excel/LibreOffice Calc")
        print("  2. Import into Tableau for advanced visualization")
        print("  3. Use pandas for programmatic analysis")
        
        print("\n📁 Files created:")
        for file_info in files_to_export:
            if file_info['excel'].exists():
                print(f"   - {file_info['excel']}")
    
    return 0 if success_count == len(files_to_export) else 1


if __name__ == "__main__":
    sys.exit(main())
